import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from copy import copy
import io
import os
import re
import datetime

st.set_page_config(page_title="Hệ thống Báo cáo Tiết dạy", page_icon="🌟", layout="wide")

st.title("Hệ thống Nộp và Tổng hợp Báo cáo")
st.write("Phiên bản 15.1: Radar quét ngang Hàng 10 để bắt tuyệt đối Số tiết Kiêm nhiệm.")

# --- CÁC DANH MỤC CỐ ĐỊNH & TỪ KHÓA NHẬN DIỆN ---
DANH_SACH_LOP = ['6A1', '6A2', '6A3', '6A4', '7A1', '7A2', '7A3', '7A4', '8A1', '8A2', '8A3', '8A4', '9A1', '9A2', '9A3', '9A4']

MON_HOC_CHINH = [
    {"tt": "1", "ten": "Ngữ văn", "loai": "don", "keys": ["văn"]},
    {"tt": "2", "ten": "Toán", "loai": "don", "keys": ["toán"]},
    {"tt": "3", "ten": "Tiếng Anh", "loai": "don", "keys": ["anh"]},
    {"tt": "4", "ten": "GDCD", "loai": "don", "keys": ["gdcd", "công dân"]},
    {"tt": "5", "ten": "LS&ĐL", "loai": "gop", "con": [
        {"ten": "Lịch sử", "keys": ["sử"]}, 
        {"ten": "Địa lí", "keys": ["địa"], "avoid": ["địa phương"]}
    ]},
    {"tt": "6", "ten": "KHTN", "loai": "gop", "con": [
        {"ten": "Lí", "keys": ["lí", "lý"], "avoid": ["địa lí", "địa lý", "quản lí", "quản lý"]}, 
        {"ten": "Hóa", "keys": ["hóa"]}, 
        {"ten": "Sinh", "keys": ["sinh"], "avoid": ["sinh hoạt", "shl"]}
    ]},
    {"tt": "7", "ten": "Công nghệ", "loai": "don", "keys": ["công nghệ"]},
    {"tt": "8", "ten": "Tin học", "loai": "don", "keys": ["tin"]},
    {"tt": "9", "ten": "GDTC", "loai": "don", "keys": ["gdtc", "thể dục"]},
    {"tt": "10", "ten": "Nghệ thuật", "loai": "gop", "con": [
        {"ten": "Mĩ thuật", "keys": ["mĩ thuật", "mỹ thuật", "mĩ", "mỹ"]}, 
        {"ten": "Âm nhạc", "keys": ["nhạc"]}
    ]},
    {"tt": "11", "ten": "HĐTN&HN", "loai": "don", "keys": ["hđtn", "trải nghiệm"]},
    {"tt": "12", "ten": "GDĐP", "loai": "don", "keys": ["gdđp", "địa phương"]},
]

def copy_sheet(source_sheet, target_sheet):
    for col, dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col].width = dim.width
    for row, dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row].height = dim.height
    for row in source_sheet.iter_rows():
        for cell in row:
            if type(cell).__name__ == 'MergedCell': continue
            target_cell = target_sheet.cell(row=cell.row, column=cell.column)
            target_cell.value = cell.value 
            if cell.has_style:
                try:
                    if cell.border: target_cell.border = copy(cell.border)
                    if cell.fill: target_cell.fill = copy(cell.fill)
                    if cell.number_format: target_cell.number_format = copy(cell.number_format)
                    if cell.alignment: target_cell.alignment = copy(cell.alignment)
                except: pass
            try:
                if cell.font:
                    new_color = copy(cell.font.color) if cell.font.color else None
                    target_cell.font = Font(name='Times New Roman', size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, vertAlign=cell.font.vertAlign, underline=cell.font.underline, strike=cell.font.strike, color=new_color)
                else:
                    target_cell.font = Font(name='Times New Roman', size=11)
            except:
                target_cell.font = Font(name='Times New Roman', size=11)
    for merge_range in source_sheet.merged_cells.ranges:
        try: target_sheet.merge_cells(str(merge_range))
        except: pass 

def get_num(sheet, row, col):
    v = sheet.cell(row=row, column=col).value
    if v is None or str(v).strip() == "": return 0
    s = str(v).strip().replace(',', '.')
    match = re.search(r'-?\d+(\.\d+)?', s)
    if match:
        val = float(match.group())
        return int(val) if val.is_integer() else val
    return 0

def is_match(val, keys, avoid=None):
    if avoid and any(a in val for a in avoid): return False
    return any(k in val for k in keys)

def get_table_bounds(sheet):
    start_row = 17 
    for r in range(1, 40):
        val_a = str(sheet.cell(row=r, column=1).value).strip().upper()
        if val_a == "TT":
            if str(sheet.cell(row=r+1, column=1).value).strip() in ["1", "01"]:
                start_row = r + 1
            else:
                start_row = r + 2
            break
            
    end_row = sheet.max_row
    for r in range(start_row, sheet.max_row + 1):
        val_b = str(sheet.cell(row=r, column=2).value or "").lower()
        val_c = str(sheet.cell(row=r, column=3).value or "").lower()
        val_a = str(sheet.cell(row=r, column=1).value or "").lower()
        if "cộng" in val_b or "tổng" in val_b or "cộng" in val_c or "tổng" in val_c or "cộng" in val_a or "tổng" in val_a:
            end_row = r - 1
            break
    return start_row, end_row

def create_program_sheet(wb_merged, list_of_sheets, nam_hoc, hoc_ky, tuan):
    ws_ct = wb_merged.create_sheet(title="Chương trình", index=0)
    bold_font = Font(name='Times New Roman', size=11, bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    prog_data = {} 
    for sheet_name in list_of_sheets:
        source_ws = wb_merged[sheet_name]
        start_row, end_row = get_table_bounds(source_ws)
        
        for r in range(start_row, end_row + 1): 
            mon_val = str(source_ws.cell(row=r, column=5).value or "").strip() 
            lop_val = str(source_ws.cell(row=r, column=6).value or "").strip().upper() 
            tiet_val = get_num(source_ws, r, 7) 

            if mon_val and lop_val in DANH_SACH_LOP:
                mon_val_lower = mon_val.lower()
                mon_found = None
                for m in MON_HOC_CHINH:
                    if mon_found: break
                    if m['loai'] == 'don':
                        if is_match(mon_val_lower, m.get('keys', []), m.get('avoid', [])): mon_found = m['ten']
                    elif m['loai'] == 'gop':
                        for c in m['con']:
                            if is_match(mon_val_lower, c.get('keys', []), c.get('avoid', [])):
                                mon_found = c['ten']; break
                if mon_found:
                    if mon_found not in prog_data: prog_data[mon_found] = {}
                    prog_data[mon_found][lop_val] = max(prog_data[mon_found].get(lop_val, 0), tiet_val)

    ws_ct.merge_cells('A1:C1'); ws_ct['A1'] = "UBND XÃ BA TƠ"; ws_ct['A1'].font = bold_font
    ws_ct.merge_cells('E1:L1'); ws_ct['E1'] = "CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM"; ws_ct['E1'].font = bold_font; ws_ct['E1'].alignment = center_align
    ws_ct.merge_cells('A2:C2'); ws_ct['A2'] = "TRƯỜNG THCS BA TƠ"; ws_ct['A2'].font = bold_font
    ws_ct.merge_cells('E2:L2'); ws_ct['E2'] = "Độc lập - Tự do - Hạnh phúc"; ws_ct['E2'].font = bold_font; ws_ct['E2'].alignment = center_align
    ws_ct.merge_cells('A4:S4'); ws_ct['A4'] = "BÁO CÁO THỰC HIỆN TIẾN ĐỘ CHƯƠNG TRÌNH"; ws_ct['A4'].font = bold_font; ws_ct['A4'].alignment = center_align
    ws_ct.merge_cells('A5:S5'); ws_ct['A5'] = f"({tuan}, {hoc_ky}, năm học {nam_hoc})"; ws_ct['A5'].font = bold_font; ws_ct['A5'].alignment = center_align
    ws_ct.merge_cells('A7:A8'); ws_ct['A7'] = "TT"
    ws_ct.merge_cells('B7:B8'); ws_ct['B7'] = "Môn/HĐGD"
    ws_ct.merge_cells('C7:R7'); ws_ct['C7'] = "Lớp"
    ws_ct.merge_cells('S7:S8'); ws_ct['S7'] = "Ghi chú"

    for c_idx, lop in enumerate(DANH_SACH_LOP, 3): ws_ct.cell(row=8, column=c_idx).value = lop
    for r in range(7, 9):
        for c in range(1, 20):
            cell = ws_ct.cell(row=r, column=c); cell.border = thin_border; cell.font = bold_font; cell.alignment = center_align

    curr_row = 9
    for m in MON_HOC_CHINH:
        ws_ct.cell(row=curr_row, column=1).value = m['tt']
        ws_ct.cell(row=curr_row, column=2).value = m['ten']
        if m['loai'] == 'gop':
            for c_idx in range(3, 19): 
                ws_ct.cell(row=curr_row, column=c_idx).value = "x"; ws_ct.cell(row=curr_row, column=c_idx).alignment = center_align
        else:
            for c_idx, lop in enumerate(DANH_SACH_LOP, 3):
                val = prog_data.get(m['ten'], {}).get(lop, "")
                if val == 0: val = ""
                ws_ct.cell(row=curr_row, column=c_idx).value = val; ws_ct.cell(row=curr_row, column=c_idx).alignment = center_align
        for c in range(1, 20): ws_ct.cell(row=curr_row, column=c).border = thin_border
        ws_ct.cell(row=curr_row, column=2).font = bold_font
        curr_row += 1

        if m['loai'] == 'gop':
            for con in m['con']:
                ws_ct.cell(row=curr_row, column=2).value = f"- {con['ten']}"
                for c_idx, lop in enumerate(DANH_SACH_LOP, 3):
                    val = prog_data.get(con['ten'], {}).get(lop, "")
                    if val == 0: val = ""
                    ws_ct.cell(row=curr_row, column=c_idx).value = val; ws_ct.cell(row=curr_row, column=c_idx).alignment = center_align
                for c in range(1, 20): ws_ct.cell(row=curr_row, column=c).border = thin_border
                curr_row += 1

    curr_row += 2
    ws_ct.merge_cells(f'M{curr_row}:S{curr_row}'); ws_ct[f'M{curr_row}'] = "Ba Tơ, ngày      tháng     năm 2026"; ws_ct[f'M{curr_row}'].alignment = center_align
    curr_row += 1
    ws_ct.merge_cells(f'M{curr_row}:S{curr_row}'); ws_ct[f'M{curr_row}'] = "PHÓ HIỆU TRƯỞNG"; ws_ct[f'M{curr_row}'].font = bold_font; ws_ct[f'M{curr_row}'].alignment = center_align
    ws_ct.column_dimensions['A'].width = 5
    ws_ct.column_dimensions['B'].width = 15
    for c in 'CDEFGHIJKLMNOPQR': ws_ct.column_dimensions[c].width = 5
    ws_ct.column_dimensions['S'].width = 25

def create_summary_sheet(wb_merged, list_of_sheets, nam_hoc, hoc_ky, tuan):
    ws_th = wb_merged.create_sheet(title="Tổng hợp")
    th_font = Font(name='Times New Roman', size=11)
    bold_font = Font(name='Times New Roman', size=11, bold=True)
    center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws_th['A1'] = "UBND XÃ BA TƠ"; ws_th['D1'] = "CỘNG HOÀ XÃ HỘI CHỦ NGHĨA VIỆT NAM"
    ws_th['A2'] = "TRƯỜNG THCS BA TƠ"; ws_th['D2'] = "Độc lập – Tự do – Hạnh phúc"
    ws_th.merge_cells('A1:B1'); ws_th.merge_cells('D1:G1'); ws_th.merge_cells('A2:B2'); ws_th.merge_cells('D2:G2')
    
    title_text = f"BẢNG TỔNG HỢP THEO DÕI BÁO CÁO THỰC HIỆN CHƯƠNG TRÌNH, TIẾT DẠY HÀNG TUẦN - NĂM HỌC {nam_hoc}"
    ws_th['A4'] = title_text; ws_th.merge_cells('A4:K4'); ws_th['A4'].alignment = center_aligned; ws_th['A4'].font = bold_font
    ws_th['A5'] = f"({hoc_ky})"; ws_th.merge_cells('A5:K5'); ws_th['A5'].alignment = center_aligned

    headers = ["TT", "Họ và tên CB, giáo viên", "Tổng số tiết thực dạy", "Số tiết kiêm nhiệm", "Số tiết đi công tác", "Số tiết dạy thay", "Số tiết lấp giờ, tăng tiết", "Số tiết coi KT, dự giờ", "Tổng số tiết thực hiện", "Số tiết Thừa/Thiếu", "Ghi chú"]
    for col, h in enumerate(headers, 1):
        cell = ws_th.cell(row=7, column=col); cell.value = h; cell.font = bold_font; cell.alignment = center_aligned; cell.border = thin_border
    
    row_idx = 8
    for tt, sheet_name in enumerate(list_of_sheets, 1):
        source_ws = wb_merged[sheet_name]
        
        start_row, end_row = get_table_bounds(source_ws)
        
        t_day = sum(get_num(source_ws, r, 8) for r in range(start_row, end_row + 1)) 
        t_cong = sum(get_num(source_ws, r, 9) for r in range(start_row, end_row + 1)) 
        t_thay = sum(get_num(source_ws, r, 10) for r in range(start_row, end_row + 1)) 
        t_tang = sum(get_num(source_ws, r, 11) for r in range(start_row, end_row + 1)) 
        t_coi = sum(get_num(source_ws, r, 12) for r in range(start_row, end_row + 1)) 
        
        # --- RADAR QUÉT NGANG HÀNG 10 TÌM SỐ KIÊM NHIỆM ---
        t_kiem = 0
        for c in range(1, 20):
            val_str = str(source_ws.cell(row=10, column=c).value or "").lower()
            if "kiêm nhiệm" in val_str:
                s_val = val_str.replace(',', '.')
                match = re.search(r'\d+(\.\d+)?', s_val)
                if match:
                    t_kiem = float(match.group())
                    break
                else:
                    for next_c in range(c + 1, 20):
                        tmp = get_num(source_ws, 10, next_c)
                        if tmp > 0: 
                            t_kiem = tmp
                            break
                    break
        
        t_kiem = int(t_kiem) if float(t_kiem).is_integer() else t_kiem
                    
        tong = t_day + t_kiem + t_cong + t_thay + t_tang + t_coi
        thua_thieu = tong - 19 
        
        vals = [tt, sheet_name, t_day, t_kiem, t_cong, t_thay, t_tang, t_coi, tong, thua_thieu, ""]
        for col, v in enumerate(vals, 1):
            cell = ws_th.cell(row=row_idx, column=col); cell.value = v; cell.border = thin_border; cell.alignment = center_aligned
        row_idx += 1

SAVE_DIR = "Du_Lieu_Bao_Cao"
if not os.path.exists(SAVE_DIR): os.makedirs(SAVE_DIR)

DANH_SACH_GV = ["Nguyễn Văn Lộc", "Đỗ Văn Linh", "Huỳnh Thị Hạ Quyên", "Phạm Thị Mỹ Thuận", "Huỳnh Thị Huyên", "Đỗ Đặng Toàn", "Bùi Thị Xuân Nhựt", "Nguyễn Thị Như Ái", "Phạm Thị Lai Tình", "Trần Văn Hoàng", "Ngô Hữu Hoá", "Phạm Thuỵ Thuỳ Nghi", "Đỗ Thanh Vũ", "Phạm Bá Quyết", "Võ Quang Tuyên", "Nguyễn Văn Thân", "Nguyễn Minh Văn", "Lê Thị Tuyết Lệ", "Trần Đình Thảo", "Bùi Thanh Tâm", "Bùi Thị Bích Vân", "K Mah Ri Lan", "Nguyễn Mẫn Thu", "Lê Thị Kim Tuyết", "Lê Thị Tường Vy", "Nguyễn Thị Thúy Hằng", "Trần Thị Kim Anh", "Nguyễn Thị Hoan", "Đinh Thị Xuân Trâm"]

st.sidebar.header("Cài đặt thông số")
nam_hoc = st.sidebar.text_input("Năm học:", "2025 - 2026")
hoc_ky = st.sidebar.selectbox("Học kỳ:", ["HỌC KỲ I", "HỌC KỲ II"], index=1)
danh_sach_tuan = [f"Tuần {i}" for i in range(1, 36)]

tab1, tab2 = st.tabs(["📤 Khu vực Giáo viên nộp bài", "⚙️ Khu vực Quản lý tổng hợp"])

with tab1:
    st.header(f"Nộp báo cáo ({hoc_ky} - {nam_hoc})")
    tuan_nop = st.selectbox("Chọn tuần báo cáo:", danh_sach_tuan, index=24) 
    ten_gv = st.selectbox("Chọn tên của bạn:", DANH_SACH_GV)
    uploaded_file = st.file_uploader("Chọn file Excel", type=['xlsx'])
    if st.button("📤 Nộp Báo Cáo"):
        if uploaded_file:
            path = os.path.join(SAVE_DIR, nam_hoc.replace(" ",""), hoc_ky.replace(" ","_"), tuan_nop.replace(" ","_"))
            if not os.path.exists(path): os.makedirs(path)
            with open(os.path.join(path, f"{ten_gv}.xlsx"), "wb") as f: f.write(uploaded_file.getbuffer())
            st.success("Nộp thành công!")

with tab2:
    st.header("Danh sách và Tổng hợp")
    mat_khau = st.text_input("🔑 Nhập mật khẩu để truy cập khu vực Quản lý:", type="password")
    
    if mat_khau == "bato2026": 
        st.success("Mở khóa thành công!")
        tuan_th = st.selectbox("Chọn tuần tổng hợp:", danh_sach_tuan, index=24)
        path_th = os.path.join(SAVE_DIR, nam_hoc.replace(" ",""), hoc_ky.replace(" ","_"), tuan_th.replace(" ","_"))
        
        files = [f for f in os.listdir(path_th) if f.endswith('.xlsx')] if os.path.exists(path_th) else []
        
        gv_da_nop_info = []
        gv_da_nop_names = []
        if files:
            for f in files:
                name = f.replace('.xlsx', '')
                gv_da_nop_names.append(name)
                mtime = os.path.getmtime(os.path.join(path_th, f))
                time_str = datetime.datetime.fromtimestamp(mtime).strftime("%H:%M - %d/%m")
                gv_da_nop_info.append(f"{name} (Lúc {time_str})")
        
        gv_chua_nop = [g for g in DANH_SACH_GV if g not in gv_da_nop_names]

        col1, col2 = st.columns(2)
        with col1: 
            st.subheader(f"✅ Đã nộp ({len(gv_da_nop_names)})")
            with st.container(height=250):
                for info in gv_da_nop_info: st.write(f"- {info}")
                
        with col2: 
            st.subheader(f"⏳ Chưa nộp ({len(gv_chua_nop)})")
            with st.container(height=250):
                for g in gv_chua_nop: st.write(f"- {g}")

        st.write("---")
        if files and st.button(f"⚙️ Tiến hành tổng hợp {tuan_th}"):
            wb_merged = openpyxl.Workbook()
            wb_merged.remove(wb_merged.active)
            with st.spinner('Đang tổng hợp dữ liệu...'):
                list_s = []
                for f in files:
                    ws_src = openpyxl.load_workbook(os.path.join(path_th, f), data_only=True).active
                    name = f.replace('.xlsx','')[:31]
                    ws_tgt = wb_merged.create_sheet(title=name)
                    copy_sheet(ws_src, ws_tgt)
                    list_s.append(name)
                
                create_program_sheet(wb_merged, list_s, nam_hoc, hoc_ky, tuan_th)
                create_summary_sheet(wb_merged, list_s, nam_hoc, hoc_ky, tuan_th)
                
                output = io.BytesIO()
                wb_merged.save(output)
                st.download_button(label="📥 Tải file tổng hợp", data=output.getvalue(), file_name=f"Tong_hop_{tuan_th}.xlsx")
    elif mat_khau != "":
        st.error("❌ Mật khẩu không chính xác! Vui lòng thử lại.")
